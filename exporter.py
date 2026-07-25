"""
Exports business records to CSV, JSON, and professionally styled Excel files.
The Excel export includes auto-fitted columns, header styling, freeze panes, auto-filters,
a summary statistics tab, and data charts.
"""

import json
import csv
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

# Openpyxl styling utilities
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from logger import logger

class Exporter:
    @staticmethod
    def export_to_csv(businesses: List[Dict[str, Any]], file_path: str) -> bool:
        """Exports businesses list to a CSV file."""
        try:
            if not businesses:
                return False
                
            df = pd.DataFrame(businesses)
            
            # Flatten opening hours and types for readable columns
            if "opening_hours" in df.columns:
                df["opening_hours"] = df["opening_hours"].apply(
                    lambda x: ", ".join(x.get("weekday_text", [])) if isinstance(x, dict) and x.get("weekday_text") else ""
                )
            if "business_types" in df.columns:
                df["business_types"] = df["business_types"].apply(
                    lambda x: ", ".join(x) if isinstance(x, list) else ""
                )
            if "accessibility" in df.columns:
                df["accessibility"] = df["accessibility"].apply(
                    lambda x: "Wheelchair Accessible" if isinstance(x, dict) and x.get("wheelchair_accessible_entrance") else "Unknown"
                )

            # Re-order/rename column headers for a clean look
            headers_mapping = {
                "name": "Business Name",
                "place_id": "Google Place ID",
                "full_address": "Full Address",
                "city": "City",
                "state": "State",
                "country": "Country",
                "postal_code": "Postal Code",
                "latitude": "Latitude",
                "longitude": "Longitude",
                "phone_number": "Phone Number",
                "international_phone_number": "International Phone Number",
                "website": "Website",
                "maps_url": "Google Maps URL",
                "rating": "Rating",
                "total_reviews": "Total Reviews",
                "business_status": "Business Status",
                "opening_hours": "Opening Hours",
                "business_types": "Business Types",
                "price_level": "Price Level",
                "plus_code": "Plus Code",
                "accessibility": "Accessibility Information",
                "created_at": "Date Collected"
            }
            
            # Select columns that exist in the dataframe
            cols_to_keep = [col for col in headers_mapping.keys() if col in df.columns]
            df = df[cols_to_keep]
            df = df.rename(columns=headers_mapping)
            
            df.to_csv(file_path, index=False, encoding="utf-8-sig")
            logger.info(f"Successfully exported {len(businesses)} records to CSV: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to export to CSV: {e}")
            return False

    @staticmethod
    def export_to_json(businesses: List[Dict[str, Any]], file_path: str) -> bool:
        """Exports businesses list to a JSON file."""
        try:
            if not businesses:
                return False
                
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(businesses, f, indent=4, ensure_ascii=False)
                
            logger.info(f"Successfully exported {len(businesses)} records to JSON: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to export to JSON: {e}")
            return False

    @staticmethod
    def export_to_excel(businesses: List[Dict[str, Any]], file_path: str) -> bool:
        """
        Exports businesses to a highly styled Excel workbook including a summary sheet,
        structured data sheet, auto-sizing, header styling, filters, frozen panes, and charts.
        """
        try:
            if not businesses:
                return False

            # Create dataframes
            df_raw = pd.DataFrame(businesses)
            
            # Flatten complex structures for Excel output
            opening_hours_col = ""
            if "opening_hours" in df_raw.columns:
                opening_hours_col = df_raw["opening_hours"].apply(
                    lambda x: "\n".join(x.get("weekday_text", [])) if isinstance(x, dict) and x.get("weekday_text") else ""
                )
            types_col = ""
            if "business_types" in df_raw.columns:
                types_col = df_raw["business_types"].apply(
                    lambda x: ", ".join(x) if isinstance(x, list) else ""
                )
            access_col = ""
            if "accessibility" in df_raw.columns:
                access_col = df_raw["accessibility"].apply(
                    lambda x: "Yes" if isinstance(x, dict) and x.get("wheelchair_accessible_entrance") is True else ("No" if isinstance(x, dict) and x.get("wheelchair_accessible_entrance") is False else "Unknown")
                )

            # Build cleaned DataFrame
            df = pd.DataFrame()
            df["Business Name"] = df_raw.get("name", "")
            df["Google Place ID"] = df_raw.get("place_id", "")
            df["Full Address"] = df_raw.get("full_address", "")
            df["City"] = df_raw.get("city", "")
            df["State"] = df_raw.get("state", "")
            df["Country"] = df_raw.get("country", "")
            df["Postal Code"] = df_raw.get("postal_code", "")
            df["Latitude"] = df_raw.get("latitude", 0.0)
            df["Longitude"] = df_raw.get("longitude", 0.0)
            df["Phone Number"] = df_raw.get("phone_number", "")
            df["International Phone Number"] = df_raw.get("international_phone_number", "")
            df["Website"] = df_raw.get("website", "")
            df["Google Maps URL"] = df_raw.get("maps_url", "")
            df["Rating"] = df_raw.get("rating", 0.0)
            df["Total Reviews"] = df_raw.get("total_reviews", 0)
            df["Business Status"] = df_raw.get("business_status", "")
            df["Opening Hours"] = opening_hours_col
            df["Business Types"] = types_col
            df["Price Level"] = df_raw.get("price_level", "")
            df["Plus Code"] = df_raw.get("plus_code", "")
            df["Wheelchair Accessible"] = access_col
            df["Date Collected"] = df_raw.get("created_at", "")

            # Open a new workbook with Openpyxl
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # --- Sheet 1: Summary Sheet ---
            ws_summary = wb.create_sheet(title="Summary & Statistics")
            ws_summary.views.sheetView[0].showGridLines = True
            
            # Stylings
            navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            soft_blue_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            bold_font = Font(name="Segoe UI", size=11, bold=True)
            regular_font = Font(name="Segoe UI", size=11)
            title_font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            # Title
            ws_summary["A1"] = "Bulk Business Search & Export Pro - Summary Report"
            ws_summary["A1"].font = title_font
            ws_summary["A2"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary["A2"].font = Font(name="Segoe UI", size=10, italic=True)

            # Statistics table headers
            stats_headers = ["Metric", "Value"]
            for col_num, header in enumerate(stats_headers, 1):
                cell = ws_summary.cell(row=4, column=col_num)
                cell.value = header
                cell.font = white_font
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="center")

            # Calculate stats
            total_biz = len(df)
            ratings = df["Rating"].dropna()
            avg_rating = round(ratings.mean(), 2) if not ratings.empty else 0.0
            
            has_website_count = df["Website"].apply(lambda x: 1 if str(x).strip() != "" and str(x) != "nan" else 0).sum()
            has_phone_count = df["Phone Number"].apply(lambda x: 1 if str(x).strip() != "" and str(x) != "nan" else 0).sum()
            
            # Stats rows
            stats_rows = [
                ("Total Businesses Exported", total_biz),
                ("Average Rating", avg_rating),
                ("Businesses with Website", f"{has_website_count} ({round(has_website_count/total_biz*100, 1)}%)" if total_biz > 0 else 0),
                ("Businesses with Phone", f"{has_phone_count} ({round(has_phone_count/total_biz*100, 1)}%)" if total_biz > 0 else 0),
            ]

            for row_idx, (metric, val) in enumerate(stats_rows, 5):
                c_m = ws_summary.cell(row=row_idx, column=1, value=metric)
                c_v = ws_summary.cell(row=row_idx, column=2, value=val)
                c_m.font = regular_font
                c_m.border = thin_border
                c_v.font = bold_font
                c_v.border = thin_border
                c_v.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

            # Rating ranges for the chart
            ws_summary.cell(row=10, column=1, value="Rating Range").font = white_font
            ws_summary.cell(row=10, column=1).fill = navy_fill
            ws_summary.cell(row=10, column=2, value="Count").font = white_font
            ws_summary.cell(row=10, column=2).fill = navy_fill

            rating_4_5 = df[df["Rating"] >= 4.5].shape[0]
            rating_4_0 = df[(df["Rating"] >= 4.0) & (df["Rating"] < 4.5)].shape[0]
            rating_3_0 = df[(df["Rating"] >= 3.0) & (df["Rating"] < 4.0)].shape[0]
            rating_low = df[(df["Rating"] < 3.0) & (df["Rating"] > 0)].shape[0]
            rating_none = df[df["Rating"].isna() | (df["Rating"] == 0)].shape[0]

            ranges = [
                ("Excellent (4.5+)", rating_4_5),
                ("Very Good (4.0-4.4)", rating_4_0),
                ("Good (3.0-3.9)", rating_3_0),
                ("Poor (<3.0)", rating_low),
                ("Unrated", rating_none)
            ]

            for r_idx, (rng, count) in enumerate(ranges, 11):
                c_rng = ws_summary.cell(row=r_idx, column=1, value=rng)
                c_cnt = ws_summary.cell(row=r_idx, column=2, value=count)
                c_rng.font = regular_font
                c_rng.border = thin_border
                c_cnt.font = regular_font
                c_cnt.border = thin_border
                c_cnt.alignment = Alignment(horizontal="right")

            # Add a beautiful bar chart representing rating ranges
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Business Rating Distribution"
            chart.y_axis.title = "Number of Businesses"
            chart.x_axis.title = "Rating Range"

            # Chart data references (row 10 to 15, column 2)
            data_ref = Reference(ws_summary, min_col=2, min_row=10, max_row=15)
            # Chart categories (row 11 to 15, column 1)
            cats_ref = Reference(ws_summary, min_col=1, min_row=11, max_row=15)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None # No legend needed for single-series column chart
            chart.width = 16
            chart.height = 10
            
            # Position chart
            ws_summary.add_chart(chart, "D4")

            # --- Sheet 2: Data Sheet ---
            ws_data = wb.create_sheet(title="Business Directory")
            ws_data.views.sheetView[0].showGridLines = True

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws_data.cell(row=1, column=col_idx, value=col_name)
                cell.font = white_font
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

            # Write business data rows
            for row_idx, row_values in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_values, 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    # Formatting values for numbers
                    if isinstance(value, float) and (col_idx in [8, 9]):  # Lat/Lng
                        cell.value = value
                        cell.number_format = '0.000000'
                    elif isinstance(value, float) and col_idx == 14:  # Rating
                        cell.value = value
                        cell.number_format = '0.0'
                    elif isinstance(value, (int, float)) and col_idx == 15:  # Reviews
                        cell.value = int(value)
                        cell.number_format = '#,##0'
                    elif isinstance(value, (int, float)) and col_idx == 19:  # Price Level
                        cell.value = int(value)
                        cell.number_format = '0'
                    else:
                        cell.value = str(value) if pd.notna(value) else ""
                        
                        # Add linebreaks inside Opening Hours to look neat
                        if col_idx == 17 and cell.value:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Apply Excel Filters
            ws_data.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

            # Freeze the top row
            ws_data.freeze_panes = "A2"

            # Auto-fit column widths
            for col in ws_data.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # We skip checking all rows if dataframe is massive to speed up
                sample_rows = col[:200] if len(col) > 200 else col
                
                for cell in sample_rows:
                    if cell.value:
                        val_str = str(cell.value)
                        # Account for linebreaks in length
                        lines = val_str.split("\n")
                        max_line_len = max(len(l) for l in lines)
                        if max_line_len > max_len:
                            max_len = max_line_len
                
                # Padding
                adjusted_width = max(max_len + 3, 11)
                # Cap column width at 50 to avoid insanely wide columns for descriptions/addresses
                ws_data.column_dimensions[col_letter].width = min(adjusted_width, 55)

            # Auto-fit Summary sheet column A
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 15

            # Save the workbook
            wb.save(file_path)
            
            logger.info(f"Successfully exported {len(businesses)} records to styled Excel sheet: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}", exc_info=True)
            return False
